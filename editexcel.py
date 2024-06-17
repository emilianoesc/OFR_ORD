from openpyxl import load_workbook
import os
def removerows(save_path,idx,amount):
    workbook = load_workbook(filename=save_path)
    #open workbook
    sheet = workbook.active
    sheet.delete_rows(idx, amount)
    #save the file
    workbook.save(filename=save_path)
        
def get_maximum_rows(save_path):
    workbook = load_workbook(filename=save_path)
    #open workbook
    sheet = workbook.active
    rows = 0
    for max_row, row in enumerate(sheet, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def edit_excel(filename,folder,rows,listcost,listname,listcolor,tipo):
    
    name2=filename.upper()
    if name2.find('SIL') !=- 1:
      cat=name2[name2.find('SIL'):name2.find('SIL')+6].strip()
    elif name2.upper().find('COP') !=- 1:
      cat=name2[name2.find('COP'):name2.find('COP')+6].strip()
    else:
      cat=name2[name2.find('SUP'):name2.find('SUP')+6].strip()



      valid_response = False
      while not valid_response:
        mat=str(input("P OR SS OR SIL:")) 
        if mat.upper() == 'P' or  mat.upper() == 'SS' or mat.upper() == 'SIL':
          valid_response = True
        else:
          print("Invalid response. Please enter P OR SS OR SIL") 
      matdict={'P':'PLATED','SS':'STAINLESS STEEL','SIL':'SILVER'}
      mat2=matdict[mat.upper()]
      cliente=str(input("CLIENTE:")).upper() 
    
    save_path = folder +"/"+ filename
    workbook = load_workbook(filename=save_path)
    #open workbook
    sheet = workbook.active
    print(rows)

    listcost2 = list(filter(None, listcost))
    listname2 = list(filter(None, listname))
    listcolor2 = list(filter(None, listcolor))

    for i, v in enumerate(listcost2):
        sheet.cell(row=i+2, column=7).value = v
    if tipo.upper()=='XO':
      for i, v in enumerate(listname2):
        sheet.cell(row=i+2, column=3).value = v
    for i, v in enumerate(listcolor2):
        sheet.cell(row=i+2, column=4).value = v + ' ' + cliente

    for row in range(2,rows+2):
      sheet.cell(row,2).value = cat
      sheet.cell(row,5).value = mat2
      sheet.cell(row,6).value = 16
      if tipo.upper()=='SO':
        sheet.cell(row,1).value = 'SPECIAL ORDERS'
        sheet.cell(row,3).value = 'ORDERS'
      else:
        sheet.cell(row,1).value = 'XIJIAO ORDERS'


    #save the file
    workbook.save(filename=save_path)
        





def out_info(save_path,colname,colcolor,colcost,colqty):
    
    workbook = load_workbook(filename=save_path)#open workbook
    sheet = workbook.active
    mylist_n = [c.value for c in sheet[colname]][1:]
    mylist_c = [c.value for c in sheet[colcost]][1:]
    mylist_q = [c.value for c in sheet[colqty]][1:]
    mylist_col = [c.value for c in sheet[colcolor]][1:] 

    mylist_c2=[s.strip('¥') if isinstance(s, str) else s for s in mylist_c]
    try:
      mylistc3=[float(s) if s is not None else '' for s in mylist_c2]
    except:
       mylistc3=mylist_c2
       print("HAY UN STRING EN LOS COSTOS")

    keys = [*range(1, len(mylist_n)+1, 1)] 

    
    lisres=[]
    for i,x in enumerate(mylist_col): 
       res = {}      
       res['color']=x
       res['name']=mylist_n[i]
       res['cost']=mylistc3[i]
       res['qty']=mylist_q[i]
       lisres.append(res)
    
    #lisresfilter=[person for person in lisres if person['qty'] != None]
    lisresfilter=lisres
    listcolor=[list(i.values())[0] for i in lisresfilter]
    listname=[list(i.values())[1] for i in lisresfilter]
    listcost=[list(i.values())[2] for i in lisresfilter]
    listqty=[list(i.values())[3] for i in lisresfilter]
    return listcolor,listname,listcost,listqty

