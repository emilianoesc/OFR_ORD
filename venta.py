from openpyxl import load_workbook
import os
import shutil
from editexcel import get_maximum_rows





def hacerventa(cell_costo='F'):
  path=r'C:\Users\MELY\Desktop'
  folder = os.path.join(path, 'COSTO A VENTA')
  ut = float(input("UTILIDAD:"))
  cell_costo_num=int(ord(cell_costo.lower())-96)

  ven='VENTA'
  precio='PRECIO'
  for  file in os.listdir(folder):
    if file[0]!='~':
    #print(file)
      filename=folder +"/"+ file
      fileventa= file.replace('COSTO',ven)
      filenameventa=folder +"/"+ fileventa
      shutil.copy(filename, filenameventa)

      workbook = load_workbook(filenameventa)
      sheet = workbook.active
      sheet.cell(1,cell_costo_num).value = precio
      #print(get_maximum_rows(sheet_object=sheet))
      mylist_c = [c.value for c in sheet[cell_costo]][1:]
      try:
        mylist_v = [round(float(i) * ut,2) if i is not None else '' for i in mylist_c]
        for i, v in enumerate(mylist_v):
          sheet.cell(row=i+2, column=cell_costo_num).value = v
        print(mylist_v)
        workbook.save(filenameventa)
      except Exception as error:
        print(error)
        os.remove(filenameventa)
      
      shutil.move(filename, path)
      shutil.move(filenameventa, path)

if __name__ == "__main__":
  hacerventa()






